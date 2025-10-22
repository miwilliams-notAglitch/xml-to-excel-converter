"""
XML to Excel Converter - Streamlit App
Save this as: app.py

To run locally (if you have Python):
  pip install streamlit openpyxl
  streamlit run app.py

To deploy to Streamlit Cloud (NO installation needed):
  See instructions below
"""

import streamlit as st
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill
import io
from datetime import datetime

# Page configuration
st.set_page_config(
    page_title="XML to Excel Converter",
    page_icon="📊",
    layout="centered"
)

# Custom CSS for better styling
st.markdown("""
    <style>
    .main {
        padding-top: 2rem;
    }
    .stButton>button {
        width: 100%;
        background-color: #667eea;
        color: white;
        font-weight: 600;
        padding: 0.75rem;
        border-radius: 0.5rem;
    }
    .stButton>button:hover {
        background-color: #764ba2;
        border-color: #764ba2;
    }
    .success-box {
        padding: 1rem;
        background-color: #d4edda;
        border-radius: 0.5rem;
        border-left: 4px solid #28a745;
        margin: 1rem 0;
    }
    .info-box {
        padding: 1rem;
        background-color: #d1ecf1;
        border-radius: 0.5rem;
        border-left: 4px solid #17a2b8;
        margin: 1rem 0;
    }
    </style>
""", unsafe_allow_html=True)

def parse_xml_to_excel(xml_file):
    """Parse XML file and create Excel output"""
    try:
        # Parse XML content
        tree = ET.parse(xml_file)
        root = tree.getroot()
        
        data_rows = []
        
        # Extract data from XML - all fields from your structure
        for transaction in root.findall('Transaction'):
            transType = transaction.find('TransType').text if transaction.find('TransType') is not None else ''
            date = transaction.find('Date').text if transaction.find('Date') is not None else ''
            time = transaction.find('Time').text if transaction.find('Time') is not None else ''
            
            # Patron information
            lastName = transaction.find('LastName').text if transaction.find('LastName') is not None else ''
            firstName = transaction.find('FirstName').text if transaction.find('FirstName') is not None else ''
            middleName = transaction.find('MiddleName').text if transaction.find('MiddleName') is not None else ''
            patronBarcode = transaction.find('PatronBarcode').text if transaction.find('PatronBarcode') is not None else ''
            districtID = transaction.find('DistrictID').text if transaction.find('DistrictID') is not None else ''
            patronType = transaction.find('PatronType').text if transaction.find('PatronType') is not None else ''
            patronGradeLevel = transaction.find('PatronGradeLevel').text if transaction.find('PatronGradeLevel') is not None else ''
            patronHomeroom = transaction.find('PatronHomeroom').text if transaction.find('PatronHomeroom') is not None else ''
            
            # Material information
            title = transaction.find('Title').text if transaction.find('Title') is not None else ''
            isbn = transaction.find('ISBN').text if transaction.find('ISBN') is not None else ''
            bibType = transaction.find('BibType').text if transaction.find('BibType') is not None else ''
            pubYear = transaction.find('PubYear').text if transaction.find('PubYear') is not None else ''
            copyBarcode = transaction.find('CopyBarcode').text if transaction.find('CopyBarcode') is not None else ''
            callNumber = transaction.find('CallNumber').text if transaction.find('CallNumber') is not None else ''
            circType = transaction.find('CircType').text if transaction.find('CircType') is not None else ''
            
            # System information
            originatorUsername = transaction.find('OriginatorUsername').text if transaction.find('OriginatorUsername') is not None else ''
            
            data_rows.append({
                'Transaction Type': transType,
                'Date': date,
                'Time': time,
                'Last Name': lastName,
                'First Name': firstName,
                'Middle Name': middleName,
                'Patron Barcode': patronBarcode,
                'District ID': districtID,
                'Patron Type': patronType,
                'Grade Level': patronGradeLevel,
                'Homeroom': patronHomeroom,
                'Title': title,
                'ISBN': isbn,
                'Material Type': bibType,
                'Publication Year': pubYear,
                'Copy Barcode': copyBarcode,
                'Call Number': callNumber,
                'Circulation Type': circType,
                'Originator Username': originatorUsername
            })
        
        if not data_rows:
            return None, "No transaction data found in XML file", 0
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Add headers with styling
        headers = ['Transaction Type', 'Date', 'Time', 'Last Name', 'First Name', 'Middle Name',
                   'Patron Barcode', 'District ID', 'Patron Type', 'Grade Level', 'Homeroom',
                   'Title', 'ISBN', 'Material Type', 'Publication Year', 'Copy Barcode', 
                   'Call Number', 'Circulation Type', 'Originator Username']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Add data rows
        for row_num, data in enumerate(data_rows, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=data[header])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes buffer
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer, "Success", len(data_rows)
        
    except ET.ParseError as e:
        return None, f"Invalid XML file: {str(e)}", 0
    except Exception as e:
        return None, f"Error processing file: {str(e)}", 0

# Main app
def main():
    # Header
    st.title("📊 XML to Excel Converter")
    st.markdown("Convert your transaction XML files to formatted Excel spreadsheets")
    
    # Info box
    st.markdown("""
        <div class="info-box">
            <strong>📋 How to use:</strong><br>
            1. Upload your XML file below<br>
            2. Click "Convert to Excel"<br>
            3. Download your formatted Excel file
        </div>
    """, unsafe_allow_html=True)
    
    # File uploader
    uploaded_file = st.file_uploader(
        "Choose an XML file",
        type=['xml'],
        help="Upload your transaction XML file (max 200MB)"
    )
    
    if uploaded_file is not None:
        # Show file details
        st.write(f"**File:** {uploaded_file.name}")
        st.write(f"**Size:** {uploaded_file.size / 1024:.2f} KB")
        
        # Convert button
        if st.button("🔄 Convert to Excel", type="primary"):
            with st.spinner("Processing your file..."):
                # Process the file
                excel_buffer, message, num_transactions = parse_xml_to_excel(uploaded_file)
                
                if excel_buffer:
                    # Success!
                    st.markdown(f"""
                        <div class="success-box">
                            <strong>✅ Success!</strong><br>
                            Processed {num_transactions} transactions
                        </div>
                    """, unsafe_allow_html=True)
                    
                    # Generate download filename
                    original_name = uploaded_file.name.rsplit('.', 1)[0]
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    download_name = f"{original_name}_transactions_{timestamp}.xlsx"
                    
                    # Download button
                    st.download_button(
                        label="📥 Download Excel File",
                        data=excel_buffer,
                        file_name=download_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                    
                    st.balloons()
                else:
                    # Error
                    st.error(f"❌ {message}")
    
    # Footer
    st.markdown("---")
    st.markdown("""
        <div style="text-align: center; color: #666; font-size: 0.9rem;">
            Upload an XML file to extract transaction data and download as Excel
        </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
